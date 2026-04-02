"""BDU FSTEC vulnerability database sync from vullist.xlsx."""

import logging
import os
import re
import ssl
import tempfile
from datetime import datetime, timezone
from typing import Any

import httpx
import psycopg2
import psycopg2.extras
from openpyxl import load_workbook

logger = logging.getLogger("red_lycoris.bdu_sync")

BDU_XLSX_URL = os.getenv(
    "BDU_XLSX_URL", "https://bdu.fstec.ru/files/documents/vullist.xlsx"
)

# Regex to extract CVE and CWE identifiers.
CVE_RE = re.compile(r"CVE-\d{4}-\d{4,}", re.IGNORECASE)
CWE_RE = re.compile(r"CWE-\d{1,5}", re.IGNORECASE)


def _cell_str(row: tuple, idx: int) -> str:
    """Safely extract a cell value as a stripped string."""
    if idx >= len(row):
        return ""
    val = row[idx]
    if val is None:
        return ""
    return str(val).strip()


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower()


def _find_header_col(header: list[str], *variants: str) -> int:
    for i, raw in enumerate(header):
        col = _normalize_header(raw)
        for variant in variants:
            if variant in col:
                return i
    return -1


def _detect_header_row(rows: list[tuple[Any, ...]]) -> int:
    scan_limit = min(len(rows), 10)
    for idx in range(scan_limit):
        row = [str(v or "").strip() for v in rows[idx]]
        has_id = _find_header_col(row, "идентификатор", "identifier", "bdu") >= 0
        has_name = _find_header_col(row, "название по", "название программного", "software name", "product name") >= 0
        if has_id and has_name:
            return idx
    return -1


def _build_vuln_column_map(header: list[str]) -> dict[str, int]:
    return {
        "bdu_id": _find_header_col(header, "идентификатор", "identifier", "bdu"),
        "name": _find_header_col(header, "наименование уязвимости", "название уязвимости", "vulnerability name"),
        "description": _find_header_col(header, "описание уязвимости", "описание", "description"),
        "vendor": _find_header_col(header, "вендор", "vendor"),
        "software_name": _find_header_col(header, "название по", "название программного", "software name", "product name"),
        "software_version": _find_header_col(header, "версия по", "software version", "version"),
        "software_type": _find_header_col(header, "тип по", "software type", "тип"),
        "os_hardware": _find_header_col(header, "наименование ос", "platform", "аппаратной платформы"),
        "vuln_class": _find_header_col(header, "класс уязвимости", "vulnerability class"),
        "detection_date": _find_header_col(header, "дата выявления", "detection date"),
        "cvss_v2": _find_header_col(header, "cvss 2", "cvss2"),
        "cvss_v3": _find_header_col(header, "cvss 3", "cvss3"),
        "cvss_v4": _find_header_col(header, "cvss 4", "cvss4"),
        "severity": _find_header_col(header, "уровень опасности", "severity"),
        "remediation": _find_header_col(header, "меры по устранению", "remediation"),
        "status": _find_header_col(header, "статус уязвимости", "status"),
        "exploit_exists": _find_header_col(header, "наличие эксплойта", "exploit"),
        "fix_info": _find_header_col(header, "информация об устранении", "fix info"),
        "source_urls": _find_header_col(header, "ссылки на источники", "references", "source"),
        "other_ids": _find_header_col(header, "идентификаторы других систем", "other id"),
        "other_info": _find_header_col(header, "прочая информация", "other info"),
        "incident_info": _find_header_col(header, "инцидентами", "incident"),
        "exploitation_method": _find_header_col(header, "способ эксплуатации", "exploitation method"),
        "fix_method": _find_header_col(header, "способ устранения", "fix method"),
        "published_date": _find_header_col(header, "дата публикации", "published date"),
        "updated_date": _find_header_col(header, "дата последнего обновления", "updated date"),
        "consequences": _find_header_col(header, "последствия эксплуатации", "consequences"),
        "vuln_state": _find_header_col(header, "состояние уязвимости", "vulnerability state"),
        "cwe_description": _find_header_col(header, "описание ошибки cwe", "cwe description"),
        "cwe_id": _find_header_col(header, "тип ошибки cwe", "cwe id"),
    }


def _build_component_column_map(header: list[str]) -> dict[str, int]:
    return {
        "bdu_id": _find_header_col(header, "идентификатор", "identifier", "bdu"),
        "vendor": _find_header_col(header, "вендор", "vendor", "производитель"),
        "software_name": _find_header_col(header, "название по", "название программного", "software name", "product name"),
        "software_version": _find_header_col(header, "версия по", "software version", "version"),
        "software_type": _find_header_col(header, "тип по", "software type", "тип"),
        "os_platform": _find_header_col(header, "наименование ос", "platform", "ос"),
    }


def _get_dsn() -> str:
    """Build PostgreSQL DSN from environment variables."""
    host = os.getenv("DB_HOST", "postgres")
    port = os.getenv("DB_PORT", "5432")
    user = os.getenv("DB_USER", "postgres")
    password = os.getenv("DB_PASSWORD", "postgres")
    dbname = os.getenv("DB_NAME", "red_lycoris")
    sslmode = os.getenv("DB_SSLMODE", "disable")
    return f"host={host} port={port} user={user} password={password} dbname={dbname} sslmode={sslmode}"


def _extract_identifiers(other_ids: str, cwe_id: str) -> list[tuple[str, str]]:
    """Extract (identifier, bdu_id) pairs from the other_ids and cwe_id fields."""
    identifiers: list[tuple[str, str]] = []
    seen: set[str] = set()

    for match in CVE_RE.findall(other_ids):
        norm = match.upper()
        if norm not in seen:
            seen.add(norm)
            identifiers.append((norm, ""))

    for match in CWE_RE.findall(cwe_id):
        norm = match.upper()
        if norm not in seen:
            seen.add(norm)
            identifiers.append((norm, ""))

    # Also extract CWE from other_ids if present there.
    for match in CWE_RE.findall(other_ids):
        norm = match.upper()
        if norm not in seen:
            seen.add(norm)
            identifiers.append((norm, ""))

    return identifiers


def download_xlsx(url: str) -> str:
    """Download the BDU xlsx file to a temporary location. Returns file path."""
    logger.info("Downloading BDU xlsx from %s", url)
    # BDU FSTEC uses a Russian CA — skip TLS verification.
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE

    with httpx.Client(verify=ctx, timeout=300, follow_redirects=True) as client:
        resp = client.get(url)
        resp.raise_for_status()

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.write(resp.content)
    tmp.close()
    logger.info("Downloaded %d bytes to %s", len(resp.content), tmp.name)
    return tmp.name


def parse_and_store(xlsx_path: str) -> dict[str, Any]:
    """Parse the xlsx file and upsert records into PostgreSQL."""
    dsn = _get_dsn()
    logger.info("Connecting to PostgreSQL")
    conn = psycopg2.connect(dsn)

    try:
        _ensure_sync_status_row(conn)
        _ensure_software_name_index(conn)
        _mark_syncing(conn, True)

        wb = load_workbook(xlsx_path, read_only=True, data_only=True)

        vuln_by_id: dict[str, tuple] = {}
        ident_set: set[tuple[str, str]] = set()
        component_set: set[tuple[str, str, str, str, str, str]] = set()
        skipped = 0

        vuln_sheet = wb["Уязвимости"] if "Уязвимости" in wb.sheetnames else wb.worksheets[0]
        vuln_rows = [row for row in vuln_sheet.iter_rows(values_only=True)]
        vuln_header_idx = _detect_header_row(vuln_rows)
        if vuln_header_idx < 0:
            raise RuntimeError("BDU vulnerabilities sheet header row not found")
        vuln_cols = _build_vuln_column_map([str(v or "").strip() for v in vuln_rows[vuln_header_idx]])

        for row in vuln_rows[vuln_header_idx + 1 :]:
            bdu_id = _cell_str(row, vuln_cols["bdu_id"])
            if not bdu_id or not bdu_id.upper().startswith("BDU:"):
                skipped += 1
                continue
            other_ids = _cell_str(row, vuln_cols["other_ids"])
            cwe_id = _cell_str(row, vuln_cols["cwe_id"])

            vuln_by_id[bdu_id] = (
                bdu_id,
                _cell_str(row, vuln_cols["name"]),
                _cell_str(row, vuln_cols["description"]),
                _cell_str(row, vuln_cols["vendor"]),
                _cell_str(row, vuln_cols["software_name"]),
                _cell_str(row, vuln_cols["software_version"]),
                _cell_str(row, vuln_cols["software_type"]),
                _cell_str(row, vuln_cols["os_hardware"]),
                _cell_str(row, vuln_cols["vuln_class"]),
                _cell_str(row, vuln_cols["detection_date"]),
                _cell_str(row, vuln_cols["cvss_v2"]),
                _cell_str(row, vuln_cols["cvss_v3"]),
                _cell_str(row, vuln_cols["cvss_v4"]),
                _cell_str(row, vuln_cols["severity"]),
                _cell_str(row, vuln_cols["remediation"]),
                _cell_str(row, vuln_cols["status"]),
                _cell_str(row, vuln_cols["exploit_exists"]),
                _cell_str(row, vuln_cols["fix_info"]),
                _cell_str(row, vuln_cols["source_urls"]),
                other_ids,
                _cell_str(row, vuln_cols["other_info"]),
                _cell_str(row, vuln_cols["incident_info"]),
                _cell_str(row, vuln_cols["exploitation_method"]),
                _cell_str(row, vuln_cols["fix_method"]),
                _cell_str(row, vuln_cols["published_date"]),
                _cell_str(row, vuln_cols["updated_date"]),
                _cell_str(row, vuln_cols["consequences"]),
                _cell_str(row, vuln_cols["vuln_state"]),
                _cell_str(row, vuln_cols["cwe_description"]),
                cwe_id,
            )
            for ident, _ in _extract_identifiers(other_ids, cwe_id):
                ident_set.add((ident, bdu_id))

        if "Компоненты" in wb.sheetnames:
            comp_rows = [row for row in wb["Компоненты"].iter_rows(values_only=True)]
            comp_header_idx = _detect_header_row(comp_rows)
            if comp_header_idx >= 0:
                comp_cols = _build_component_column_map([str(v or "").strip() for v in comp_rows[comp_header_idx]])
                for row in comp_rows[comp_header_idx + 1 :]:
                    bdu_id = _cell_str(row, comp_cols["bdu_id"])
                    software_name = _cell_str(row, comp_cols["software_name"])
                    software_version = _cell_str(row, comp_cols["software_version"])
                    if not bdu_id or not bdu_id.upper().startswith("BDU:") or not software_name:
                        continue
                    component_set.add((
                        bdu_id,
                        _cell_str(row, comp_cols["vendor"]),
                        software_name,
                        software_version,
                        _cell_str(row, comp_cols["software_type"]),
                        _cell_str(row, comp_cols["os_platform"]),
                    ))
            else:
                logger.warning("Компоненты: header row not found, skipping components import")

        wb.close()

        _replace_dataset(conn, list(vuln_by_id.values()), list(ident_set), list(component_set))

        # Commit the entire sync atomically — either all rows land or none.
        conn.commit()

        total = len(vuln_by_id)
        _update_sync_status(conn, total, None)
        logger.info(
            "BDU sync complete: %d vulnerabilities, %d components, %d skipped",
            total,
            len(component_set),
            skipped,
        )
        return {"status": "completed", "records": total, "components": len(component_set), "skipped": skipped}

    except Exception as exc:
        logger.exception("BDU sync failed")
        conn.rollback()
        try:
            _update_sync_status(conn, 0, str(exc))
        except Exception:
            pass
        raise
    finally:
        conn.close()


def _replace_dataset(
    conn,
    vulns: list[tuple],
    idents: list[tuple[str, str]],
    components: list[tuple[str, str, str, str, str, str]],
) -> None:
    """Replace vulnerabilities/components dataset and identifier mappings."""
    with conn.cursor() as cur:
        cur.execute("DELETE FROM bdu_identifier_map")
        cur.execute("DELETE FROM bdu_vulnerabilities")
        cur.execute("DELETE FROM bdu_components")

        # Upsert vulnerabilities.
        if vulns:
            psycopg2.extras.execute_values(
                cur,
                """
                INSERT INTO bdu_vulnerabilities (
                    bdu_id, name, description, vendor, software_name, software_version,
                    software_type, os_hardware, vuln_class, detection_date,
                    cvss_v2, cvss_v3, cvss_v4, severity, remediation, status,
                    exploit_exists, fix_info, source_urls, other_ids, other_info,
                    incident_info, exploitation_method, fix_method, published_date,
                    updated_date, consequences, vuln_state, cwe_description, cwe_id,
                    synced_at
                ) VALUES %s
                ON CONFLICT (bdu_id) DO NOTHING
                """,
                [(v + (datetime.now(timezone.utc),)) for v in vulns],
                page_size=1000,
            )

        # Upsert identifier mappings.
        if idents:
            psycopg2.extras.execute_values(
                cur,
                """
                INSERT INTO bdu_identifier_map (identifier, bdu_id)
                VALUES %s
                ON CONFLICT (identifier, bdu_id) DO NOTHING
                """,
                idents,
                page_size=2000,
            )

        if components:
            psycopg2.extras.execute_values(
                cur,
                """
                INSERT INTO bdu_components (
                    bdu_id, vendor, software_name, software_version, software_type, os_platform, created_at
                ) VALUES %s
                ON CONFLICT ON CONSTRAINT uq_bdu_comp DO NOTHING
                """,
                [(c + (datetime.now(timezone.utc),)) for c in components],
                page_size=2000,
            )


def _mark_syncing(conn, syncing: bool) -> None:
    with conn.cursor() as cur:
        cur.execute(
            "UPDATE bdu_sync_status SET is_syncing = %s, updated_at = NOW() WHERE id = 1",
            (syncing,),
        )
    conn.commit()


def _update_sync_status(conn, count: int, error: str | None) -> None:
    with conn.cursor() as cur:
        cur.execute(
            """
            UPDATE bdu_sync_status
            SET last_synced_at = NOW(),
                record_count = %s,
                last_error = %s,
                is_syncing = FALSE,
                updated_at = NOW()
            WHERE id = 1
            """,
            (count, error),
        )
    conn.commit()


def _ensure_sync_status_row(conn) -> None:
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO bdu_sync_status (id, sync_interval_hours, record_count, is_syncing, updated_at)
            VALUES (1, 24, 0, FALSE, NOW())
            ON CONFLICT (id) DO NOTHING
            """
        )
    conn.commit()


def _ensure_software_name_index(conn) -> None:
    with conn.cursor() as cur:
        cur.execute("DROP INDEX IF EXISTS public.idx_bdu_vulnerabilities_software_name_lower")
        cur.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_bdu_vulnerabilities_software_name_md5
            ON public.bdu_vulnerabilities (md5(LOWER(software_name)))
            """
        )
    conn.commit()
